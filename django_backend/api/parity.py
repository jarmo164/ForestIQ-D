"""Main-branch commercial and operational workflows implemented for Django REST."""

from __future__ import annotations

import csv
import hashlib
import io
import json
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation

from django.conf import settings
from django.db import transaction
from django.db.models import Max
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from accounts.models import PrivilegeCode, User
from forestry.models import Cadastre, DataSyncRun, InheritanceSignal, Owner, OwnerLog
from forestry.tasks import enqueue_cadastre_sync, enqueue_portfolio_sync
from operations.models import (
    Contract,
    ContractHistory,
    Deal,
    DealOffer,
    DealStage,
    InheritanceCase,
    InheritanceCaseEvent,
    InheritanceHeir,
    OwnerImportBatch,
    OwnershipTransitionEvent,
    Reminder,
)

from .permissions import CanEvaluate, CanManageOwners, IsAdmin, can_access_owner
from .organization import organization_user_or_404, organization_users, request_organization_id
from .serializers import cadastre_summary, json_value, owner_summary, user_data


def _detail(message: str, http_status: int = status.HTTP_400_BAD_REQUEST) -> Response:
    return Response({"detail": message}, status=http_status)


def _number(value, field: str, *, required: bool = False) -> Decimal | None:
    if value in (None, ""):
        if required:
            raise ValueError(f"{field} is required.")
        return None
    try:
        parsed = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"{field} must be a valid decimal value.") from exc
    if parsed < 0:
        raise ValueError(f"{field} cannot be negative.")
    return parsed


def _date(value, field: str) -> date | None:
    if value in (None, ""):
        return None
    try:
        return date.fromisoformat(str(value))
    except ValueError as exc:
        raise ValueError(f"{field} must use YYYY-MM-DD.") from exc


def _owner_access(request, owner_id: str):
    owner = get_object_or_404(Owner.objects.prefetch_related("cadastres").select_related("assignee"), id=owner_id)
    if not can_access_owner(request.user, owner):
        return None, _detail("You do not have access to this owner.", status.HTTP_403_FORBIDDEN)
    return owner, None


def _deal_data(deal: Deal) -> dict:
    offers = [
        {
            "id": str(item.id), "revision": item.revision, "kind": item.kind, "status": item.status,
            "amount": json_value(item.amount), "validUntil": item.valid_until.isoformat() if item.valid_until else None,
            "terms": item.terms, "note": item.note, "createdBy": user_data(item.created_by),
            "createdAt": json_value(item.created_at), "sentAt": json_value(item.sent_at),
            "acceptedAt": json_value(item.accepted_at), "rejectedAt": json_value(item.rejected_at),
        }
        for item in deal.offers.select_related("created_by").all()
    ]
    return {
        "id": str(deal.id), "owner": owner_summary(deal.owner), "saleSubject": deal.sale_subject,
        "stage": deal.stage, "parcels": [cadastre_summary(item) for item in deal.parcels.all()],
        "decisionMaker": deal.decision_maker or None, "saleTimeframe": deal.sale_timeframe or None,
        "priceExpectation": json_value(deal.price_expectation), "qualificationNotes": deal.qualification_notes or None,
        "evaluator": user_data(deal.evaluator), "evaluationStatus": deal.evaluation_status or None,
        "estimatedMinPrice": json_value(deal.estimated_min_price), "estimatedMaxPrice": json_value(deal.estimated_max_price),
        "recommendedPurchasePrice": json_value(deal.recommended_purchase_price), "internalMinPrice": json_value(deal.internal_min_price),
        "proposedOfferPrice": json_value(deal.proposed_offer_price),
        "offerValidUntil": deal.offer_valid_until.isoformat() if deal.offer_valid_until else None,
        "assumptions": deal.evaluation_assumptions or None, "risks": deal.evaluation_risks or None,
        "returnedReason": deal.returned_reason or None, "lossReason": deal.loss_reason or None,
        "closedAt": json_value(deal.closed_at), "createdAt": json_value(deal.created_at), "updatedAt": json_value(deal.updated_at),
        "offers": offers,
    }


def _get_deal(deal_id: str) -> Deal:
    return get_object_or_404(
        Deal.objects.select_related("owner", "evaluator", "created_by").prefetch_related("parcels", "offers__created_by"), id=deal_id
    )


def _ensure_deal_access(request, deal: Deal):
    if not can_access_owner(request.user, deal.owner):
        return _detail("You do not have access to this deal.", status.HTTP_403_FORBIDDEN)
    return None


@api_view(["GET"])
@permission_classes([CanManageOwners])
def deal_brief(request, owner_id: str):
    owner, denied = _owner_access(request, owner_id)
    if denied:
        return denied
    return Response({"owner": owner_summary(owner), "parcels": [cadastre_summary(item) for item in owner.cadastres.all()], "openDeals": owner.deals.exclude(stage__in=[DealStage.WON, DealStage.LOST, DealStage.CANCELLED]).count()})


@api_view(["GET", "POST"])
@permission_classes([CanManageOwners])
def deals_by_owner(request, owner_id: str):
    owner, denied = _owner_access(request, owner_id)
    if denied:
        return denied
    if request.method == "GET":
        deals = owner.deals.select_related("owner", "evaluator", "created_by").prefetch_related("parcels", "offers__created_by")
        return Response([_deal_data(item) for item in deals])
    data = request.data
    sale_subject = str(data.get("saleSubject", "")).upper()
    if sale_subject not in {"FOREST", "LAND", "BOTH"}:
        return _detail("saleSubject must be FOREST, LAND or BOTH.")
    parcel_ids = data.get("parcelIds", data.get("parcels", []))
    if not isinstance(parcel_ids, list) or not parcel_ids:
        return _detail("At least one parcelId is required.")
    parcels = list(owner.cadastres.filter(id__in=parcel_ids))
    if len(parcels) != len(set(parcel_ids)):
        return _detail("Every selected parcel must belong to the owner.")
    try:
        price_expectation = _number(data.get("priceExpectation"), "priceExpectation")
    except ValueError as exc:
        return _detail(str(exc))
    evaluator = None
    evaluator_id = data.get("evaluatorId") or data.get("evaluatorSubject")
    if evaluator_id:
        evaluator = organization_user_or_404(request, evaluator_id, active_only=True)
    request_evaluation = bool(data.get("requestEvaluation", False))
    with transaction.atomic():
        deal = Deal.objects.create(
            owner=owner, sale_subject=sale_subject, stage=DealStage.EVALUATION if request_evaluation else DealStage.QUALIFICATION,
            decision_maker=str(data.get("decisionMaker", "")), sale_timeframe=str(data.get("saleTimeframe", "")),
            price_expectation=price_expectation, qualification_notes=str(data.get("qualificationNotes", "")),
            evaluator=evaluator, created_by=request.user,
        )
        deal.parcels.set(parcels)
        OwnerLog.objects.create(owner=owner, creator=request.user, message=f"Created commercial deal {deal.id} ({sale_subject}).")
    return Response(_deal_data(_get_deal(str(deal.id))), status=status.HTTP_201_CREATED)


@api_view(["GET"])
@permission_classes([CanEvaluate])
def deal_evaluation_queue(request):
    records = Deal.objects.filter(stage=DealStage.EVALUATION).select_related("owner", "evaluator").prefetch_related("parcels", "offers")
    if not request.user.has_privilege(PrivilegeCode.ADMIN) and not request.user.has_privilege(PrivilegeCode.OWNER_PROFILE):
        records = records.filter(evaluator__in=[request.user, None])
    return Response([_deal_data(item) for item in records])


@api_view(["POST", "DELETE"])
@permission_classes([CanEvaluate])
def deal_evaluation_claim(request, deal_id: str):
    deal = _get_deal(deal_id)
    denied = _ensure_deal_access(request, deal)
    if denied:
        return denied
    if request.method == "POST":
        if deal.evaluator_id and deal.evaluator_id != request.user.id and not request.user.has_privilege(PrivilegeCode.ADMIN):
            return _detail("The evaluation is already assigned.", status.HTTP_409_CONFLICT)
        deal.evaluator = request.user
        deal.stage = DealStage.EVALUATION
        deal.save(update_fields=["evaluator", "stage", "updated_at"])
    elif deal.evaluator_id == request.user.id or request.user.has_privilege(PrivilegeCode.ADMIN):
        deal.evaluator = None
        deal.save(update_fields=["evaluator", "updated_at"])
    else:
        return _detail("Only the assigned evaluator can release the case.", status.HTTP_403_FORBIDDEN)
    return Response(_deal_data(_get_deal(deal_id)))


@api_view(["PUT"])
@permission_classes([IsAdmin])
def deal_evaluation_assignment(request, deal_id: str):
    deal = _get_deal(deal_id)
    evaluator_id = request.data.get("evaluatorId") or request.data.get("evaluatorSubject")
    deal.evaluator = organization_user_or_404(request, evaluator_id, active_only=True) if evaluator_id else None
    deal.stage = DealStage.EVALUATION
    deal.save(update_fields=["evaluator", "stage", "updated_at"])
    return Response(_deal_data(_get_deal(deal_id)))


@api_view(["POST"])
@permission_classes([CanEvaluate])
def deal_evaluation_submit(request, deal_id: str):
    deal = _get_deal(deal_id)
    denied = _ensure_deal_access(request, deal)
    if denied:
        return denied
    if deal.evaluator_id and deal.evaluator_id != request.user.id and not request.user.has_privilege(PrivilegeCode.ADMIN):
        return _detail("Only the assigned evaluator can submit this evaluation.", status.HTTP_403_FORBIDDEN)
    evaluation_status = str(request.data.get("status", "")).upper()
    if evaluation_status not in {"NEEDS_DATA", "SUBMITTED", "APPROVED"}:
        return _detail("status must be NEEDS_DATA, SUBMITTED or APPROVED.")
    try:
        numbers = {field: _number(request.data.get(camel), camel) for field, camel in {
            "estimated_min_price": "estimatedMinPrice", "estimated_max_price": "estimatedMaxPrice",
            "recommended_purchase_price": "recommendedPurchasePrice", "internal_min_price": "internalMinPrice",
            "proposed_offer_price": "proposedOfferPrice",
        }.items()}
        offer_valid_until = _date(request.data.get("offerValidUntil"), "offerValidUntil")
    except ValueError as exc:
        return _detail(str(exc))
    deal.evaluator = deal.evaluator or request.user
    deal.evaluation_status = evaluation_status
    deal.stage = DealStage.NEGOTIATION if evaluation_status == "APPROVED" else DealStage.EVALUATION
    deal.offer_valid_until = offer_valid_until
    deal.evaluation_assumptions = str(request.data.get("assumptions", ""))
    deal.evaluation_risks = str(request.data.get("risks", ""))
    deal.returned_reason = str(request.data.get("returnedReason", ""))
    for field, value in numbers.items():
        setattr(deal, field, value)
    deal.save()
    OwnerLog.objects.create(owner=deal.owner, creator=request.user, message=f"Evaluation for deal {deal.id}: {evaluation_status}.")
    return Response(_deal_data(_get_deal(deal_id)))


def _commercial(deal: Deal) -> dict:
    data = _deal_data(deal)
    return {"dealId": data["id"], "stage": data["stage"], "nextAction": "Close or update the latest offer.", "closedAt": data["closedAt"], "offers": data["offers"]}


def _new_offer(deal: Deal, request, *, kind: str, amount: Decimal, valid_until: date | None = None, terms: str = "", note: str = "") -> DealOffer:
    revision = (deal.offers.aggregate(last=Max("revision"))["last"] or 0) + 1
    return DealOffer.objects.create(deal=deal, revision=revision, kind=kind, amount=amount, valid_until=valid_until, terms=terms, note=note, created_by=request.user)


@api_view(["GET"])
@permission_classes([CanManageOwners])
def deal_commercial(request, deal_id: str):
    deal = _get_deal(deal_id)
    denied = _ensure_deal_access(request, deal)
    return denied or Response(_commercial(deal))


@api_view(["POST"])
@permission_classes([CanManageOwners])
def deal_offer_create(request, deal_id: str):
    deal = _get_deal(deal_id)
    denied = _ensure_deal_access(request, deal)
    if denied:
        return denied
    if deal.stage in {DealStage.WON, DealStage.LOST, DealStage.CANCELLED}:
        return _detail("A closed deal cannot receive new offers.", status.HTTP_409_CONFLICT)
    try:
        offer = _new_offer(deal, request, kind=DealOffer.Kind.OFFER, amount=_number(request.data.get("amount"), "amount", required=True), valid_until=_date(request.data.get("validUntil"), "validUntil"), terms=str(request.data.get("terms", "")))
    except ValueError as exc:
        return _detail(str(exc))
    return Response({"offer": _commercial(_get_deal(deal_id))["offers"][-1], "state": _commercial(_get_deal(deal_id))}, status=status.HTTP_201_CREATED)


@api_view(["POST"])
@permission_classes([CanManageOwners])
def deal_offer_send(request, deal_id: str):
    deal = _get_deal(deal_id)
    denied = _ensure_deal_access(request, deal)
    if denied:
        return denied
    offer = get_object_or_404(DealOffer, id=request.data.get("offerId"), deal=deal)
    if offer.status != DealOffer.Status.DRAFT:
        return _detail("Only draft offers can be sent.", status.HTTP_409_CONFLICT)
    offer.status, offer.sent_at = DealOffer.Status.SENT, timezone.now()
    offer.save(update_fields=["status", "sent_at"])
    deal.stage = DealStage.NEGOTIATION
    deal.save(update_fields=["stage", "updated_at"])
    OwnerLog.objects.create(owner=deal.owner, creator=request.user, message=f"Sent offer revision {offer.revision} for deal {deal.id}.")
    return Response(_commercial(_get_deal(deal_id)))


@api_view(["POST"])
@permission_classes([CanManageOwners])
def deal_counteroffer(request, deal_id: str):
    deal = _get_deal(deal_id)
    denied = _ensure_deal_access(request, deal)
    if denied:
        return denied
    try:
        _new_offer(deal, request, kind=DealOffer.Kind.COUNTEROFFER, amount=_number(request.data.get("amount"), "amount", required=True), note=str(request.data.get("note", "")))
    except ValueError as exc:
        return _detail(str(exc))
    deal.stage = DealStage.NEGOTIATION
    deal.save(update_fields=["stage", "updated_at"])
    return Response(_commercial(_get_deal(deal_id)), status=status.HTTP_201_CREATED)


def _close_deal(deal: Deal, *, stage: str, request, message: str, loss_reason: str = "", accepted_offer: DealOffer | None = None) -> Response:
    now = timezone.now()
    deal.stage, deal.closed_at, deal.loss_reason = stage, now, loss_reason
    deal.save(update_fields=["stage", "closed_at", "loss_reason", "updated_at"])
    if accepted_offer:
        accepted_offer.status, accepted_offer.accepted_at = DealOffer.Status.ACCEPTED, now
        accepted_offer.save(update_fields=["status", "accepted_at"])
        DealOffer.objects.filter(deal=deal).exclude(id=accepted_offer.id).filter(status__in=[DealOffer.Status.DRAFT, DealOffer.Status.SENT]).update(status=DealOffer.Status.REJECTED, rejected_at=now)
    OwnerLog.objects.create(owner=deal.owner, creator=request.user, message=message[:4000])
    return Response(_commercial(_get_deal(str(deal.id))))


@api_view(["POST"])
@permission_classes([CanManageOwners])
def deal_won(request, deal_id: str):
    deal = _get_deal(deal_id)
    denied = _ensure_deal_access(request, deal)
    if denied:
        return denied
    offer = get_object_or_404(DealOffer, id=request.data.get("acceptedEntryId") or request.data.get("offerId"), deal=deal)
    return _close_deal(deal, stage=DealStage.WON, request=request, accepted_offer=offer, message=str(request.data.get("note", "Deal won.")))


@api_view(["POST"])
@permission_classes([CanManageOwners])
def deal_lost(request, deal_id: str):
    deal = _get_deal(deal_id)
    denied = _ensure_deal_access(request, deal)
    if denied:
        return denied
    reason = str(request.data.get("lossReason", "")).strip()
    if not reason:
        return _detail("lossReason is required.")
    return _close_deal(deal, stage=DealStage.LOST, request=request, loss_reason=reason, message=str(request.data.get("note", f"Deal lost: {reason}.")))


@api_view(["POST"])
@permission_classes([CanManageOwners])
def deal_cancelled(request, deal_id: str):
    deal = _get_deal(deal_id)
    denied = _ensure_deal_access(request, deal)
    if denied:
        return denied
    note = str(request.data.get("note", "")).strip()
    if not note:
        return _detail("note is required.")
    return _close_deal(deal, stage=DealStage.CANCELLED, request=request, message=note)


def _event_data(event: InheritanceCaseEvent) -> dict:
    return {"id": str(event.id), "type": event.type, "description": event.description, "createdAt": json_value(event.created_at), "createdBy": user_data(event.created_by)}


def _heir_data(heir: InheritanceHeir) -> dict:
    return {"id": str(heir.id), "displayName": heir.display_name, "personalCode": heir.personal_code or None, "registryCode": heir.registry_code or None, "inheritanceShare": heir.inheritance_share or None, "relationToDeceased": heir.relation_to_deceased or None, "phone": heir.phone or None, "email": heir.email or None, "contactStatus": heir.contact_status or None, "assignedTo": user_data(heir.assigned_to), "source": heir.source or None}


def _case_data(case: InheritanceCase) -> dict:
    return {"id": str(case.id), "owner": owner_summary(case.owner), "sourceNoticeNumber": case.source_notice_number or None, "sourceUrl": case.source_url or None, "announcementDate": case.announcement_date.isoformat() if case.announcement_date else None, "deathDate": case.death_date.isoformat() if case.death_date else None, "certificationDeadline": case.certification_deadline.isoformat() if case.certification_deadline else None, "notaryName": case.notary_name or None, "notaryPhone": case.notary_phone or None, "status": case.status, "assignedTo": user_data(case.assigned_to), "startedAt": json_value(case.started_at), "endedAt": json_value(case.ended_at), "updatedAt": json_value(case.updated_at), "heirs": [_heir_data(item) for item in case.heirs.select_related("assigned_to").all()], "events": [_event_data(item) for item in case.events.select_related("created_by").all()]}


def _get_case(case_id: str) -> InheritanceCase:
    return get_object_or_404(InheritanceCase.objects.select_related("owner", "assigned_to").prefetch_related("heirs__assigned_to", "events__created_by"), id=case_id)


def _case_access(request, case: InheritanceCase):
    if not can_access_owner(request.user, case.owner):
        return _detail("You do not have access to this inheritance case.", status.HTTP_403_FORBIDDEN)
    return None


def _parse_case_fields(data: dict) -> dict:
    try:
        return {"source_notice_number": str(data.get("sourceNoticeNumber", "")), "source_url": str(data.get("sourceUrl", "")), "announcement_date": _date(data.get("announcementDate"), "announcementDate"), "death_date": _date(data.get("deathDate"), "deathDate"), "certification_deadline": _date(data.get("certificationDeadline"), "certificationDeadline"), "notary_name": str(data.get("notaryName", "")), "notary_phone": str(data.get("notaryPhone", ""))}
    except ValueError as exc:
        raise ValueError(str(exc)) from exc


@api_view(["GET", "POST"])
@permission_classes([CanManageOwners])
def inheritance_by_owner(request, owner_id: str):
    owner, denied = _owner_access(request, owner_id)
    if denied:
        return denied
    if request.method == "GET":
        return Response([_case_data(item) for item in owner.inheritance_cases.select_related("owner", "assigned_to").prefetch_related("heirs__assigned_to", "events__created_by")])
    try:
        fields = _parse_case_fields(request.data)
    except ValueError as exc:
        return _detail(str(exc))
    case = InheritanceCase.objects.create(owner=owner, **fields)
    InheritanceCaseEvent.objects.create(inheritance_case=case, type="CREATED", description="Inheritance case created manually.", created_by=request.user)
    return Response(_case_data(_get_case(str(case.id))), status=status.HTTP_201_CREATED)


@api_view(["GET"])
@permission_classes([CanManageOwners])
def inheritance_cases(request):
    records = InheritanceCase.objects.select_related("owner", "assigned_to").prefetch_related("heirs", "events")
    if not request.user.has_privilege(PrivilegeCode.ADMIN, PrivilegeCode.OWNER_PROFILE):
        records = records.filter(assigned_to__in=[request.user, None])
    requested_status = request.query_params.get("status")
    if requested_status:
        records = records.filter(status=requested_status)
    return Response([_case_data(item) for item in records[:200]])


@api_view(["GET"])
@permission_classes([CanManageOwners])
def inheritance_case_detail(request, case_id: str):
    case = _get_case(case_id)
    denied = _case_access(request, case)
    return denied or Response(_case_data(case))


@api_view(["PATCH"])
@permission_classes([CanManageOwners])
def inheritance_case_assignment(request, case_id: str):
    case = _get_case(case_id)
    denied = _case_access(request, case)
    if denied:
        return denied
    user_id = request.data.get("assignedTo") or request.data.get("assignedToSubject")
    case.assigned_to = get_object_or_404(User, id=user_id, is_active=True) if user_id else None
    case.save(update_fields=["assigned_to", "updated_at"])
    InheritanceCaseEvent.objects.create(inheritance_case=case, type="ASSIGNMENT", description="Case assignment updated.", created_by=request.user)
    return Response(_case_data(_get_case(case_id)))


@api_view(["PATCH"])
@permission_classes([CanManageOwners])
def inheritance_case_status(request, case_id: str):
    case = _get_case(case_id)
    denied = _case_access(request, case)
    if denied:
        return denied
    new_status = str(request.data.get("status", "")).upper()
    if new_status not in InheritanceCase.Status.values:
        return _detail(f"status must be one of: {', '.join(InheritanceCase.Status.values)}.")
    case.status = new_status
    if new_status in {InheritanceCase.Status.COMPLETED, InheritanceCase.Status.CLOSED}:
        case.ended_at = timezone.now()
    else:
        case.ended_at = None
    case.save(update_fields=["status", "ended_at", "updated_at"])
    comment = str(request.data.get("comment", "")).strip() or f"Case status changed to {new_status}."
    InheritanceCaseEvent.objects.create(inheritance_case=case, type="STATUS", description=comment, created_by=request.user)
    return Response(_case_data(_get_case(case_id)))


def _save_heir(case: InheritanceCase, data: dict, *, heir: InheritanceHeir | None = None) -> InheritanceHeir:
    assignee_id = data.get("assignedTo") or data.get("assignedToSubject")
    target = heir or InheritanceHeir(inheritance_case=case)
    target.display_name = str(data.get("displayName", "")).strip()
    if not target.display_name:
        raise ValueError("displayName is required.")
    target.personal_code = str(data.get("personalCode", "")); target.registry_code = str(data.get("registryCode", ""))
    target.inheritance_share = str(data.get("inheritanceShare", "")); target.relation_to_deceased = str(data.get("relationToDeceased", ""))
    target.phone = str(data.get("phone", "")); target.email = str(data.get("email", "")); target.contact_status = str(data.get("contactStatus", ""))
    target.source = str(data.get("source", "")); target.assigned_to = organization_user_or_404(request, assignee_id, active_only=True) if assignee_id else None
    target.save()
    return target


@api_view(["POST"])
@permission_classes([CanManageOwners])
def inheritance_case_heirs(request, case_id: str):
    case = _get_case(case_id)
    denied = _case_access(request, case)
    if denied:
        return denied
    try:
        heir = _save_heir(case, request.data)
    except ValueError as exc:
        return _detail(str(exc))
    InheritanceCaseEvent.objects.create(inheritance_case=case, type="HEIR", description=f"Added heir {heir.display_name}.", created_by=request.user)
    return Response(_heir_data(heir), status=status.HTTP_201_CREATED)


@api_view(["PUT"])
@permission_classes([CanManageOwners])
def inheritance_case_heir_detail(request, case_id: str, heir_id: str):
    case = _get_case(case_id)
    denied = _case_access(request, case)
    if denied:
        return denied
    heir = get_object_or_404(InheritanceHeir, id=heir_id, inheritance_case=case)
    try:
        heir = _save_heir(case, request.data, heir=heir)
    except ValueError as exc:
        return _detail(str(exc))
    InheritanceCaseEvent.objects.create(inheritance_case=case, type="HEIR", description=f"Updated heir {heir.display_name}.", created_by=request.user)
    return Response(_heir_data(heir))


@api_view(["POST"])
@permission_classes([CanManageOwners])
def inheritance_case_event(request, case_id: str):
    case = _get_case(case_id)
    denied = _case_access(request, case)
    if denied:
        return denied
    content = str(request.data.get("content", request.data.get("description", ""))).strip()
    if not content:
        return _detail("content is required.")
    event = InheritanceCaseEvent.objects.create(inheritance_case=case, type="NOTE", description=content, created_by=request.user)
    return Response(_event_data(event), status=status.HTTP_201_CREATED)


def _notice_to_case(owner: Owner, signal: InheritanceSignal, user) -> InheritanceCase:
    case, created = InheritanceCase.objects.get_or_create(owner=owner, source_notice_number=signal.source_notice_number, defaults={"source_url": signal.source_url, "announcement_date": signal.announcement_date, "certification_deadline": signal.certification_deadline, "notary_name": "", "notary_phone": ""})
    if created:
        InheritanceCaseEvent.objects.create(inheritance_case=case, type="OFFICIAL_NOTICE", description=f"Imported official notice {signal.source_notice_number}.", created_by=user)
    return case


@api_view(["POST"])
@permission_classes([CanManageOwners])
def inheritance_official_check(request, owner_id: str):
    owner, denied = _owner_access(request, owner_id)
    if denied:
        return denied
    signal = owner.inheritance_signals.order_by("-announcement_date", "-id").first()
    if not signal:
        return Response({"sourceAvailable": bool(settings.PARIMUS_API_URL), "noticeFound": False, "message": "No imported official notice was found.", "inheritanceCase": None})
    case = _notice_to_case(owner, signal, request.user)
    return Response({"sourceAvailable": bool(settings.PARIMUS_API_URL), "noticeFound": True, "message": "Official notice is available.", "inheritanceCase": _case_data(_get_case(str(case.id)))})


@api_view(["POST"])
@permission_classes([CanManageOwners])
def inheritance_notice_import(request, owner_id: str, notice_number: str):
    owner, denied = _owner_access(request, owner_id)
    if denied:
        return denied
    signal = get_object_or_404(InheritanceSignal, owner=owner, source_notice_number=notice_number)
    case = _notice_to_case(owner, signal, request.user)
    return Response(_case_data(_get_case(str(case.id))), status=status.HTTP_201_CREATED)


IMPORT_FIELDS = ("id", "name", "type", "phone", "email", "address", "info")
IMPORT_ALIASES = {"id": ("id", "isikukood", "registry code", "registrikood", "code"), "name": ("name", "nimi", "owner name"), "type": ("type", "liik", "owner type"), "phone": ("phone", "telefon", "tel"), "email": ("email", "e-post", "epost"), "address": ("address", "aadress"), "info": ("info", "notes", "märkus", "markus")}


def _normalized(value: str) -> str:
    return " ".join(str(value).strip().lower().replace("_", " ").split())


def _read_import(uploaded) -> tuple[list[str], list[dict[str, str]], str]:
    raw = uploaded.read()
    uploaded.seek(0)
    checksum = hashlib.sha256(raw).hexdigest()
    filename = uploaded.name or "owners-import"
    if filename.lower().endswith(".csv"):
        text = raw.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return list(reader.fieldnames or []), [{key: value or "" for key, value in row.items()} for row in reader], checksum
    if filename.lower().endswith(".xlsx"):
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return [], [], checksum
        headers = [str(value or "").strip() for value in values[0]]
        rows = [{headers[index]: str(value or "").strip() for index, value in enumerate(row) if index < len(headers)} for row in values[1:] if any(value not in (None, "") for value in row)]
        return headers, rows, checksum
    raise ValueError("Only UTF-8 CSV and XLSX files are accepted.")


def _mapping(headers: list[str], supplied) -> dict[str, str]:
    values = supplied or {}
    if isinstance(values, str):
        values = json.loads(values)
    if not isinstance(values, dict):
        raise ValueError("mapping must be an object.")
    available = {_normalized(header): header for header in headers}
    mapping: dict[str, str] = {}
    for field in IMPORT_FIELDS:
        candidate = values.get(field)
        if candidate and _normalized(candidate) in available:
            mapping[field] = available[_normalized(candidate)]
    if not mapping:
        for field, aliases in IMPORT_ALIASES.items():
            candidate = next((available[_normalized(alias)] for alias in aliases if _normalized(alias) in available), None)
            if candidate:
                mapping[field] = candidate
    if not mapping.get("id") or not mapping.get("name"):
        raise ValueError("mapping must include columns for id and name.")
    if len(set(mapping.values())) != len(mapping):
        raise ValueError("One source column cannot map to multiple owner fields.")
    return mapping


def _preview_import(uploaded, supplied_mapping) -> tuple[str, str, dict[str, str], list[dict]]:
    headers, rows, checksum = _read_import(uploaded)
    mapping = _mapping(headers, supplied_mapping)
    results = []
    seen: set[str] = set()
    existing = set(Owner.objects.filter(id__in=[str(row.get(mapping["id"], "")).strip() for row in rows]).values_list("id", flat=True))
    for row_number, row in enumerate(rows, start=2):
        values = {field: str(row.get(header, "")).strip() for field, header in mapping.items()}
        owner_id, name = values.get("id", ""), values.get("name", "")
        reason, row_status = "", "READY"
        if not owner_id or not name:
            row_status, reason = "INVALID", "id and name are required"
        elif owner_id in seen or owner_id in existing:
            row_status, reason = "DUPLICATE", "owner id already exists in the file or database"
        seen.add(owner_id)
        results.append({"rowNumber": row_number, "id": owner_id, "name": name, "status": row_status, "reason": reason, "values": values})
    return uploaded.name or "owners-import", checksum, mapping, results


@api_view(["POST"])
@permission_classes([IsAdmin])
def owner_import_inspect(request):
    uploaded = request.FILES.get("file")
    if not uploaded:
        return _detail("file is required.")
    try:
        headers, _rows, checksum = _read_import(uploaded)
        suggested = _mapping(headers, {})
    except (ValueError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        return _detail(str(exc))
    return Response({"filename": uploaded.name, "sha256": checksum, "headers": headers, "suggestedMapping": suggested, "requiredFields": ["id", "name"]})


@api_view(["POST"])
@permission_classes([IsAdmin])
def owner_import_preview(request):
    uploaded = request.FILES.get("file")
    if not uploaded:
        return _detail("file is required.")
    try:
        filename, checksum, mapping, rows = _preview_import(uploaded, request.data.get("mapping"))
    except (ValueError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        return _detail(str(exc))
    return Response({"filename": filename, "sha256": checksum, "mapping": mapping, "readyCount": sum(row["status"] == "READY" for row in rows), "rejectedCount": sum(row["status"] != "READY" for row in rows), "rows": rows})


@api_view(["POST"])
@permission_classes([IsAdmin])
def owner_import_commit(request):
    uploaded = request.FILES.get("file")
    if not uploaded:
        return _detail("file is required.")
    try:
        filename, checksum, _mapping_result, rows = _preview_import(uploaded, request.data.get("mapping"))
    except (ValueError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        return _detail(str(exc))
    if checksum != str(request.data.get("confirmedSha256", "")):
        return _detail("confirmedSha256 does not match the inspected file.", status.HTTP_412_PRECONDITION_FAILED)
    rejected = [{key: value for key, value in row.items() if key != "values"} for row in rows if row["status"] != "READY"]
    created_ids: list[str] = []
    with transaction.atomic():
        for row in rows:
            if row["status"] != "READY":
                continue
            values = row["values"]
            owner = Owner.objects.create(id=values["id"], name=values["name"], type=values.get("type") or "PERSON", phone=values.get("phone", ""), email=values.get("email", ""), address=values.get("address", ""), info=values.get("info", ""))
            created_ids.append(owner.id)
        batch = OwnerImportBatch.objects.create(filename=filename, sha256=checksum, creator=request.user, created_count=len(created_ids), rejected_rows=rejected)
    return Response({"batchId": str(batch.id), "filename": filename, "sha256": checksum, "createdCount": len(created_ids), "createdOwnerIds": created_ids, "rejectedRows": rejected}, status=status.HTTP_201_CREATED)


@api_view(["GET"])
@permission_classes([IsAdmin])
def owner_import_batches(request):
    return Response([{"batchId": str(item.id), "filename": item.filename, "sha256": item.sha256, "createdBy": user_data(item.creator), "committedAt": json_value(item.committed_at), "createdCount": item.created_count, "rejectedCount": len(item.rejected_rows)} for item in OwnerImportBatch.objects.select_related("creator").all()[:100]])


@api_view(["GET"])
@permission_classes([IsAdmin])
def owner_import_errors_csv(request, batch_id: str):
    batch = get_object_or_404(OwnerImportBatch, id=batch_id)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=["rowNumber", "id", "name", "status", "reason"])
    writer.writeheader(); writer.writerows(batch.rejected_rows)
    response = HttpResponse(output.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="owner-import-{batch.id}-errors.csv"'
    return response


SALES_OUTCOMES = {"CALLBACK", "NO_ANSWER", "INTERESTED", "NOT_INTERESTED", "DECLINED", "FOLLOW_UP"}


@api_view(["GET"])
@permission_classes([CanManageOwners])
def sales_queue(request):
    records = Owner.objects.filter(assignee=request.user).select_related("assignee").prefetch_related("cadastres").order_by("status_set_at", "id")
    return Response([{"owner": owner_summary(item), "markedParcelCount": item.cadastres.filter(marked=True).count(), "openDealCount": item.deals.exclude(stage__in=[DealStage.WON, DealStage.LOST, DealStage.CANCELLED]).count(), "nextReminder": json_value(item.reminders.filter(due_time__gte=timezone.now()).order_by("due_time").values_list("due_time", flat=True).first())} for item in records])


@api_view(["POST"])
@permission_classes([CanManageOwners])
def sales_outcome(request, owner_id: str):
    owner, denied = _owner_access(request, owner_id)
    if denied:
        return denied
    outcome = str(request.data.get("outcome", "")).upper()
    if outcome not in SALES_OUTCOMES:
        return _detail(f"outcome must be one of: {', '.join(sorted(SALES_OUTCOMES))}.")
    OwnerLog.objects.create(owner=owner, creator=request.user, message=f"Sales outcome: {outcome}. {str(request.data.get('note', '')).strip()}".strip())
    reminder = None
    if outcome in {"CALLBACK", "NO_ANSWER", "FOLLOW_UP"}:
        delay = 1 if outcome == "CALLBACK" else 2
        reminder = Reminder.objects.create(owner=owner, creator=request.user, text=f"Sales follow-up: {outcome.lower().replace('_', ' ')}", due_time=timezone.now() + timedelta(days=delay))
    return Response({"owner": owner_summary(owner), "outcome": outcome, "reminder": {"id": str(reminder.id), "dueTime": json_value(reminder.due_time)} if reminder else None})


@api_view(["POST"])
@permission_classes([IsAdmin])
def ownership_transition_sync(request):
    return _detail("Forestek ownership import is a one-time initial import and cannot be started from the recurring sync API.", status.HTTP_409_CONFLICT)


@api_view(["GET"])
@permission_classes([CanManageOwners])
def ownership_transition_events(request, owner_id: str):
    owner, denied = _owner_access(request, owner_id)
    if denied:
        return denied
    return Response([{"id": str(item.id), "cadastreId": item.cadastre_id, "type": item.event_type, "occurredAt": json_value(item.occurred_at), "sourceReference": item.source_reference or None, "payload": item.payload, "recordedAt": json_value(item.recorded_at)} for item in owner.ownership_transitions.select_related("cadastre").all()])


def _integration_rows() -> list[dict]:
    rows = []
    for key, label, configured in [("CADASTRE", "Cadastre and forest registry", True), ("FORESTEK", "Forestek ownership relations", bool(settings.FORESTEK_API_URL and settings.FORESTEK_API_TOKEN)), ("PARIMUS", "Pärimus official notices", bool(settings.PARIMUS_API_URL and settings.PARIMUS_API_TOKEN))]:
        latest = DataSyncRun.objects.filter(source__icontains=key.lower()).order_by("-id").first()
        rows.append({"key": key, "label": label, "configured": configured, "mode": "ONE_TIME" if key == "FORESTEK" else "RECURRING", "lastRun": {"id": latest.id, "status": latest.status, "finishedAt": json_value(latest.finished_at), "error": latest.error_message or None} if latest else None})
    return rows


@api_view(["GET"])
@permission_classes([IsAdmin])
def integration_jobs(request):
    return Response(_integration_rows())


@api_view(["GET", "POST"])
@permission_classes([IsAdmin])
def integration_start(request, key: str):
    key = key.upper()
    if key not in {"CADASTRE", "FORESTEK", "PARIMUS"}:
        return _detail("Unknown integration key.", status.HTTP_404_NOT_FOUND)
    if request.method == "GET":
        limit = min(max(int(request.query_params.get("limit", "10")), 1), 100)
        records = DataSyncRun.objects.filter(source__icontains=key.lower()).order_by("-id")[:limit]
        return Response([{"id": item.id, "cadastreId": item.cadastre_id, "source": item.source, "status": item.status, "taskId": item.task_id, "startedAt": json_value(item.started_at), "finishedAt": json_value(item.finished_at), "result": item.result, "error": item.error_message or None} for item in records])
    if key == "FORESTEK":
        return _detail("Forestek is a one-time initial import. Run the controlled management command only before its first successful import.", status.HTTP_409_CONFLICT)
    cadastre_id = request.data.get("cadastreId") or request.data.get("parameters", {}).get("cadastreId")
    if cadastre_id:
        run = enqueue_cadastre_sync(
            str(cadastre_id),
            organization_id=str(request_organization_id(request)),
            requested_by_id=request.user.id,
            source=key.lower(),
        )
        return Response({"key": key, "status": run.status, "runId": run.id, "taskId": run.task_id}, status=status.HTTP_202_ACCEPTED)
    if settings.FORESTIQ_TASKS_INLINE:
        result = enqueue_portfolio_sync(str(request_organization_id(request)))
        return Response({"key": key, "status": "SUCCEEDED", "result": result})
    task = enqueue_portfolio_sync.delay(str(request_organization_id(request)))
    return Response({"key": key, "status": "QUEUED", "taskId": task.id}, status=status.HTTP_202_ACCEPTED)


@api_view(["GET"])
@permission_classes([IsAdmin])
def integration_runs(request, key: str):
    limit = min(max(int(request.query_params.get("limit", "10")), 1), 100)
    records = DataSyncRun.objects.filter(source__icontains=key.lower()).order_by("-id")[:limit]
    return Response([{"id": item.id, "cadastreId": item.cadastre_id, "source": item.source, "status": item.status, "taskId": item.task_id, "startedAt": json_value(item.started_at), "finishedAt": json_value(item.finished_at), "result": item.result, "error": item.error_message or None} for item in records])


@api_view(["GET"])
@permission_classes([IsAdmin])
def integration_run(request, run_id: int):
    item = get_object_or_404(DataSyncRun, id=run_id)
    return Response({"id": item.id, "cadastreId": item.cadastre_id, "source": item.source, "status": item.status, "taskId": item.task_id, "startedAt": json_value(item.started_at), "finishedAt": json_value(item.finished_at), "result": item.result, "error": item.error_message or None})


def _stale_cadastres(days: int = 30):
    threshold = timezone.now() - timedelta(days=days)
    fresh_ids = set(DataSyncRun.objects.filter(status=DataSyncRun.Status.SUCCEEDED, finished_at__gte=threshold).values_list("cadastre_id", flat=True))
    return Cadastre.objects.exclude(id__in=fresh_ids)


@api_view(["GET"])
@permission_classes([IsAdmin])
def registry_freshness(request):
    stale = _stale_cadastres()
    return Response({"totalCadastres": Cadastre.objects.count(), "staleCadastres": stale.count(), "recentFailures": DataSyncRun.objects.filter(status=DataSyncRun.Status.FAILED).count(), "oldestStaleCadastres": list(stale.order_by("id").values_list("id", flat=True)[:50])})


@api_view(["POST"])
@permission_classes([IsAdmin])
def registry_recover(request):
    batch_size = min(max(int(request.data.get("batchSize", 25)), 1), 200)
    queued = []
    for cadastre_id in _stale_cadastres().order_by("id").values_list("id", flat=True)[:batch_size]:
        queued.append(
            enqueue_cadastre_sync(
                cadastre_id,
                organization_id=str(request_organization_id(request)),
                requested_by_id=request.user.id,
                source="recovery",
            ).id
        )
    return Response({"queued": len(queued), "runIds": queued}, status=status.HTTP_202_ACCEPTED)


@api_view(["GET"])
@permission_classes([IsAdmin])
def registry_health(request):
    latest = DataSyncRun.objects.order_by("-id").first()
    return Response({"status": "DEGRADED" if DataSyncRun.objects.filter(status=DataSyncRun.Status.FAILED).exists() else "OK", "latestRun": {"id": latest.id, "status": latest.status, "finishedAt": json_value(latest.finished_at)} if latest else None, "freshness": {"staleCadastres": _stale_cadastres().count()}})


def _registry_refresh(request, cadastre_id: str, source: str):
    get_object_or_404(Cadastre, id=cadastre_id)
    run = enqueue_cadastre_sync(
        cadastre_id,
        organization_id=str(request_organization_id(request)),
        requested_by_id=request.user.id,
        source=source,
    )
    return Response({"id": run.id, "cadastreId": cadastre_id, "source": source, "status": run.status, "taskId": run.task_id}, status=status.HTTP_202_ACCEPTED)


@api_view(["POST"])
@permission_classes([IsAdmin])
def registry_refresh_all(request, cadastre_id: str):
    return _registry_refresh(request, cadastre_id, "registry_all")


@api_view(["POST"])
@permission_classes([IsAdmin])
def registry_refresh_maaamet(request, cadastre_id: str):
    return _registry_refresh(request, cadastre_id, "maaamet")


@api_view(["POST"])
@permission_classes([IsAdmin])
def registry_refresh_notifications(request, cadastre_id: str):
    return _registry_refresh(request, cadastre_id, "metsaregister_notifications")


@api_view(["POST"])
@permission_classes([IsAdmin])
def registry_refresh_plan(request, cadastre_id: str):
    return _registry_refresh(request, cadastre_id, "metsaregister_plan")


@api_view(["POST"])
@permission_classes([IsAdmin])
def registry_rik_cadastre_count(request):
    code = str(request.data.get("personalOrRegistryCode", "")).strip()
    name = " ".join(filter(None, [str(request.data.get("firstName", "")).strip(), str(request.data.get("lastNameOrCompany", "")).strip()])).strip()
    records = Owner.objects.all()
    if code:
        records = records.filter(id=code)
    elif name:
        records = records.filter(name__iexact=name)
    return Response({"cadastreCount": Cadastre.objects.filter(owners__in=records).distinct().count()})


@api_view(["GET"])
@permission_classes([IsAdmin])
def portfolio_status(request):
    latest = DataSyncRun.objects.filter(source__icontains="forestek").order_by("-id").first()
    return Response({"configured": bool(settings.FORESTEK_API_URL and settings.FORESTEK_API_TOKEN), "mode": "ONE_TIME_INITIAL_IMPORT", "initialImportCompleted": bool(latest and latest.status == DataSyncRun.Status.SUCCEEDED), "cadastreCount": Cadastre.objects.count(), "latestRun": {"id": latest.id, "status": latest.status, "finishedAt": json_value(latest.finished_at), "error": latest.error_message or None} if latest else None})


@api_view(["POST"])
@permission_classes([IsAdmin])
def portfolio_sync(request):
    return _detail("Forestek portfolio sync is disabled after the one-time initial import.", status.HTTP_409_CONFLICT)


def _deal_contract_draft(deal: Deal) -> dict:
    accepted_offer = deal.offers.filter(status=DealOffer.Status.ACCEPTED).first()
    if deal.stage != DealStage.WON or accepted_offer is None:
        raise ValueError("A deal needs an accepted offer before a contract can be drafted.")
    parcels = [{"cadastralCode": parcel.id, "address": parcel.address, "areaHectares": json_value(parcel.area)} for parcel in deal.parcels.all()]
    if not parcels:
        raise ValueError("The deal has no parcels to include in a contract.")
    return {"dealId": str(deal.id), "offerEntryId": str(accepted_offer.id), "acceptedPrice": json_value(accepted_offer.amount), "acceptedTerms": accepted_offer.terms, "seller": {"name": deal.owner.name, "code": deal.owner.id, "address": deal.owner.address, "iban": ""}, "parcels": parcels}


@api_view(["GET"])
@permission_classes([IsAdmin])
def contract_deal_draft(request, deal_id: str):
    deal = _get_deal(deal_id)
    try:
        return Response(_deal_contract_draft(deal))
    except ValueError as exc:
        return _detail(str(exc), status.HTTP_412_PRECONDITION_FAILED)


@api_view(["POST"])
@permission_classes([IsAdmin])
def contract_generate_from_deal(request):
    deal = _get_deal(str(request.data.get("dealId", "")))
    try:
        draft = _deal_contract_draft(deal)
    except ValueError as exc:
        return _detail(str(exc), status.HTTP_412_PRECONDITION_FAILED)
    contract_id = str(request.data.get("contractId") or deal.id)
    contract_number = str(request.data.get("contractNumber", "")).strip()
    if not contract_number:
        return _detail("contractNumber is required.")
    with transaction.atomic():
        history = ContractHistory.objects.create(
            id=contract_id,
            sellers=draft["seller"]["name"], buyer=str(request.data.get("buyer", "ForestIQ buyer")),
            contract_number=contract_number, created_at=timezone.now(),
            data={"deal": draft, "terms": request.data.get("terms", draft["acceptedTerms"]), "contractNumber": contract_number},
            cadastres=", ".join(item["cadastralCode"] for item in draft["parcels"]),
        )
        contract, _ = Contract.objects.update_or_create(id=contract_id, defaults={"base_id": str(request.data.get("baseId", "")), "source_deal": deal, "source_offer_id": draft["offerEntryId"]})
    return Response({"contractId": contract.id, "historyId": history.id, "dealId": str(deal.id), "offerEntryId": draft["offerEntryId"], "pdf": f"/api/services/contracts/{contract.id}/pdf"}, status=status.HTTP_201_CREATED)
